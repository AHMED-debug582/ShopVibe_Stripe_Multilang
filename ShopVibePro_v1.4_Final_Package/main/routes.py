from flask import render_template, current_app, request, redirect, session, url_for, abort, flash
from . import main
from bson.objectid import ObjectId
from werkzeug.security import generate_password_hash, check_password_hash
import stripe

# الصفحة الرئيسية
@main.route("/")
def index():
    return "ShopVibe Pro - Ready"

# عرض جميع المنتجات
@main.route("/products")
def products():
    products_collection = current_app.mongo["products"]
    products = list(products_collection.find())
    return render_template("products.html", products=products)

# عرض تفاصيل منتج واحد
@main.route("/product/<product_id>")
def product_detail(product_id):
    products_collection = current_app.mongo["products"]
    try:
        product = products_collection.find_one({"_id": ObjectId(product_id)})
    except:
        abort(404)
    if not product:
        abort(404)
    return render_template("product_detail.html", product=product)

# إضافة إلى السلة
@main.route("/add-to-cart", methods=["POST"])
def add_to_cart():
    product_id = request.form.get("product_id")
    quantity = int(request.form.get("quantity", 1))

    if "cart" not in session:
        session["cart"] = {}

    if product_id in session["cart"]:
        session["cart"][product_id] += quantity
    else:
        session["cart"][product_id] = quantity

    session.modified = True
    save_cart_to_db()
    return redirect(url_for("main.cart"))

# عرض السلة
@main.route("/cart")
def cart():
    cart = session.get("cart", {})
    products_collection = current_app.mongo["products"]

    cart_items = []
    total_price = 0

    for product_id, quantity in cart.items():
        try:
            product = products_collection.find_one({"_id": ObjectId(product_id)})
            if product:
                product["quantity"] = quantity
                product["subtotal"] = product["price"] * quantity
                total_price += product["subtotal"]
                cart_items.append(product)
        except:
            continue

    return render_template("cart.html", cart_items=cart_items, total_price=total_price)

# حذف منتج من السلة
@main.route("/remove-from-cart/<product_id>")
def remove_from_cart(product_id):
    if "cart" in session and product_id in session["cart"]:
        session["cart"].pop(product_id)
        session.modified = True
        save_cart_to_db()
    return redirect(url_for("main.cart"))

# تعديل كمية منتج في السلة
@main.route("/update-cart", methods=["POST"])
def update_cart():
    product_id = request.form.get("product_id")
    try:
        quantity = int(request.form.get("quantity"))
        if quantity < 1:
            return redirect(url_for("main.remove_from_cart", product_id=product_id))
    except:
        quantity = 1

    if "cart" in session and product_id in session["cart"]:
        session["cart"][product_id] = quantity
        session.modified = True
        save_cart_to_db()

    return redirect(url_for("main.cart"))

# صفحة الدفع Checkout
@main.route("/checkout")
def checkout():
    cart = session.get("cart", {})
    products_collection = current_app.mongo["products"]

    cart_items = []
    total_price = 0

    for product_id, quantity in cart.items():
        try:
            product = products_collection.find_one({"_id": ObjectId(product_id)})
            if product:
                product["quantity"] = quantity
                product["subtotal"] = product["price"] * quantity
                total_price += product["subtotal"]
                cart_items.append(product)
        except:
            continue

    return render_template("checkout.html", cart_items=cart_items, total_price=total_price)

# إنشاء جلسة دفع Stripe
@main.route("/create-checkout-session", methods=["POST"])
def create_checkout_session():
    cart = session.get("cart", {})
    products_collection = current_app.mongo["products"]
    line_items = []

    for product_id, quantity in cart.items():
        try:
            product = products_collection.find_one({"_id": ObjectId(product_id)})
            if product:
                line_items.append({
                    "price_data": {
                        "currency": "mad",
                        "product_data": {
                            "name": product["name"],
                        },
                        "unit_amount": int(product["price"] * 100),
                    },
                    "quantity": quantity,
                })
        except:
            continue

    if not line_items:
        return redirect(url_for("main.cart"))

    session_obj = stripe.checkout.Session.create(
        payment_method_types=["card"],
        line_items=line_items,
        mode="payment",
        success_url=url_for("main.success", _external=True),
        cancel_url=url_for("main.cancel", _external=True),
    )

    return redirect(session_obj.url, code=303)

# صفحة نجاح الدفع
from datetime import datetime

@main.route("/success")
def success():
    if "user_id" in session:
        user_id = session["user_id"]
        cart = session.get("cart", {})
        products_collection = current_app.mongo["products"]
        orders_collection = current_app.mongo["orders"]
        order_items = []
        total = 0

        for product_id, quantity in cart.items():
            try:
                product = products_collection.find_one({"_id": ObjectId(product_id)})
                if product:
                    subtotal = product["price"] * quantity
                    order_items.append({
                        "product_id": str(product["_id"]),
                        "name": product["name"],
                        "price": product["price"],
                        "quantity": quantity,
                        "subtotal": subtotal
                    })
                    total += subtotal
            except:
                continue

        orders_collection.insert_one({
            "user_id": ObjectId(user_id),
            "items": order_items,
            "total": total,
            "created_at": datetime.utcnow()
        })

    session.pop("cart", None)
    flash("✅ تم الدفع وحفظ الطلب بنجاح.")
    return redirect(url_for("main.my_account"))

# صفحة إلغاء الدفع
@main.route("/cancel")
def cancel():
    flash("❌ تم إلغاء عملية الدفع.")
    return redirect(url_for("main.cart"))

# تسجيل مستخدم جديد
@main.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        users = current_app.mongo["users"]
        if users.find_one({"email": email}):
            flash("📧 البريد الإلكتروني مستخدم بالفعل.")
            return redirect("/register")

        hashed_password = generate_password_hash(password)
        users.insert_one({"email": email, "password": hashed_password})
        flash("✅ تم التسجيل بنجاح! يمكنك تسجيل الدخول الآن.")
        return redirect("/login")

    return render_template("register.html")

# تسجيل الدخول
@main.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        users = current_app.mongo["users"]
        user = users.find_one({"email": email})
        if not user or not check_password_hash(user["password"], password):
            flash("❌ معلومات غير صحيحة.")
            return redirect("/login")

        session["user_id"] = str(user["_id"])

        # تحميل السلة الخاصة بالمستخدم
        user_cart = user.get("cart", {})
        session["cart"] = user_cart

        flash("👋 مرحبًا بعودتك!")
        return redirect("/")

    return render_template("login.html")

# تسجيل الخروج
@main.route("/logout")
def logout():
    session.pop("user_id", None)
    flash("🚪 تم تسجيل الخروج بنجاح.")
    return redirect("/")

# حفظ السلة في MongoDB
def save_cart_to_db():
    if "user_id" in session:
        user_id = session["user_id"]
        cart = session.get("cart", {})
        users = current_app.mongo["users"]
        users.update_one({"_id": ObjectId(user_id)}, {"$set": {"cart": cart}})
@main.route("/my-account")
def my_account():
    if "user_id" not in session:
        flash("🔐 يجب تسجيل الدخول أولاً.")
        return redirect("/login")

    orders_collection = current_app.mongo["orders"]
    orders = list(orders_collection.find({"user_id": ObjectId(session["user_id"])}).sort("created_at", -1))
    return render_template("my_account.html", orders=orders)
from reportlab.pdfgen import canvas
from io import BytesIO
from flask import send_file

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet

@main.route("/invoice/<order_id>")
def invoice(order_id):
    if "user_id" not in session:
        flash("🔐 يجب تسجيل الدخول.")
        return redirect("/login")

    orders = current_app.mongo["orders"]
    order = orders.find_one({"_id": ObjectId(order_id), "user_id": ObjectId(session["user_id"])})

    if not order:
        flash("❌ لا يمكن العثور على هذا الطلب.")
        return redirect("/my-account")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=60, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    # ✅ شعار
    logo_path = "static/logo.png"
    try:
        elements.append(Image(logo_path, width=80, height=50))
    except:
        pass

    elements.append(Spacer(1, 12))

    # ✅ عنوان الفاتورة
    elements.append(Paragraph(f"<b>فاتورة الطلب #{order['_id']}</b>", styles["Title"]))
    elements.append(Paragraph(f"التاريخ: {order['created_at'].strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # ✅ جدول المنتجات
    data = [["المنتج", "الكمية", "السعر", "الإجمالي"]]
    for item in order["items"]:
        data.append([
            item["name"],
            str(item["quantity"]),
            f"{item['price']} MAD",
            f"{item['subtotal']} MAD"
        ])
    data.append(["", "", "الإجمالي", f"{order['total']} MAD"])

    table = Table(data, colWidths=[150, 70, 80, 80])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"invoice_{order_id}.pdf", mimetype="application/pdf")
@main.route("/admin")
def admin_dashboard():
    if not session.get("admin_logged_in"):
        return redirect("/admin-login")

    return render_template("admin/dashboard.html")
from flask import Blueprint, render_template, request, redirect, session, flash, url_for
import hashlib

@main.route("/admin-login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        hashed = hashlib.sha256(password.encode()).hexdigest()

        admin = current_app.mongo["admins"].find_one({
            "username": username,
            "password": hashed
        })

        if admin:
            session["admin_logged_in"] = True
            return redirect(url_for("main.admin_dashboard"))
        else:
            flash("اسم المستخدم أو كلمة المرور غير صحيحة ❌")

    return render_template("admin/login.html")



@main.route("/dashboard")
def dashboard():
    db = current_app.mongo
    total_orders = db["orders"].count_documents({})
    total_users = db["users"].count_documents({})
    total_products = db["products"].count_documents({})
    total_revenue = sum(order.get("total", 0) for order in db["orders"].find({}))
    
    return render_template("dashboard.html",
                           total_orders=total_orders,
                           total_users=total_users,
                           total_products=total_products,
                           total_revenue=total_revenue)


# 📊 Dashboard + Export routes

@main.route("/dashboard")
def dashboard():
    db = current_app.mongo
    orders = db["orders"]
    users = db["users"]
    products = db["products"]

    total_orders = orders.count_documents({})
    total_users = users.count_documents({})
    total_products = products.count_documents({})
    total_revenue = sum(order.get("total", 0) for order in orders.find({}))

    from collections import defaultdict
    import calendar

    monthly_sales = defaultdict(float)
    for order in orders.find({}):
        created = order.get("created_at")
        if created:
            month = created.strftime("%b")
            monthly_sales[month] += float(order.get("total", 0))

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    chart_labels = months
    chart_data = [round(monthly_sales.get(m, 0), 2) for m in months]

    return render_template("dashboard.html",
                           total_orders=total_orders,
                           total_users=total_users,
                           total_products=total_products,
                           total_revenue=total_revenue,
                           chart_labels=chart_labels,
                           chart_data=chart_data)


@main.route("/export_excel")
def export_excel():
    from openpyxl import Workbook
    from flask import make_response
    db = current_app.mongo
    orders = db["orders"]
    users = db["users"]
    products = db["products"]

    total_orders = orders.count_documents({})
    total_users = users.count_documents({})
    total_products = products.count_documents({})
    total_revenue = sum(order.get("total", 0) for order in orders.find({}))

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Stats"
    ws.append(["Statistic", "Value"])
    ws.append(["Total Orders", total_orders])
    ws.append(["Total Users", total_users])
    ws.append(["Total Products", total_products])
    ws.append(["Total Revenue", total_revenue])

    response = make_response()
    wb.save(response.stream)
    response.headers["Content-Disposition"] = "attachment; filename=dashboard_stats.xlsx"
    response.mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@main.route("/export_pdf")
def export_pdf():
    from flask import make_response
    from fpdf import FPDF
    db = current_app.mongo
    orders = db["orders"]
    users = db["users"]
    products = db["products"]

    total_orders = orders.count_documents({})
    total_users = users.count_documents({})
    total_products = products.count_documents({})
    total_revenue = sum(order.get("total", 0) for order in orders.find({}))

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=14)
    pdf.cell(200, 10, txt="📊 ShopVibe - Dashboard Report", ln=True, align="C")
    pdf.ln(10)
    pdf.cell(100, 10, txt=f"Total Orders: {total_orders}", ln=True)
    pdf.cell(100, 10, txt=f"Total Users: {total_users}", ln=True)
    pdf.cell(100, 10, txt=f"Total Products: {total_products}", ln=True)
    pdf.cell(100, 10, txt=f"Total Revenue: {total_revenue} MAD", ln=True)

    response = make_response(pdf.output(dest='S').encode('latin1'))
    response.headers["Content-Disposition"] = "attachment; filename=dashboard_report.pdf"
    response.mimetype = "application/pdf"
    return response
